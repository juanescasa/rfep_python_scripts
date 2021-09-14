# -*- coding: utf-8 -*-
"""
Created on Wed Apr 28 11:03:54 2021

@author: calle
"""
#import all what you need
#%%
import pandas as pd
import gurobipy as gp
from gurobipy import GRB
import os
import time
import openpyxl
import sys
import jecs_functions
#this reads all data for the problem
from read_data_rfep import *
import rfep_model


#%%
#this dict stores all the tuples that output the solve of the subproblem
d_subproblem_solution = {}
d_subproblem_status = {}
d_subproblem_ovInventory = {}
d_subproblem_ovRefuelQuantity = {}
d_subproblem_ovRefuel = {}
d_subproblem_oTotalRefuellingCost = {}
d_subproblem_oTotalCost = {}

#solution statistics
d_subproblem_n_constraints = {}
d_subproblem_n_variables = {}
d_subproblem_n_integer_variables = {}
d_subproblem_n_binary_variables = {}
d_subproblem_model_fingerprint = {}
d_subproblem_model_runtime = {}
d_subproblem_model_MIPGap = {}


output_file = os.path.join("..", "output", "outputRFEPSubproblem_v1.xlsx")
#output_file = "C:\OneDrive - Deakin University\OD\calle test\Disun Applications\Gurobi Applications\output\outputRFEPSubproblem.xlsx"
workbook = openpyxl.load_workbook(output_file)
sheet_name = 'oNodesNodesVehiclesPaths'
ws = workbook[sheet_name]


for element in sVehiclesPaths:
    sSubVehiclesPaths = element
    #update sets to filter per vehicle path
    sSubNodesVehiclesPaths = [(i,v,p) for (i,v,p) in sNodesVehiclesPaths if v == element[0] and p==element[1]]
    sSubStationsVehiclesPaths = [(i,v,p) for (i,v,p) in sStationsVehiclesPaths if v == element[0] and p==element[1]]
    sSubOriginVehiclesPaths = [(i,v,p) for (i,v,p) in sOriginVehiclesPaths if v == element[0] and p==element[1]]
    sSubDestinationVehiclesPaths = [(i,v,p) for (i,v,p) in sDestinationVehiclesPaths if v == element[0] and p==element[1]]
    sSubSequenceNodesNodesVehiclesPaths = [(i,j,v,p) for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if v == element[0] and p==element[1]]
    sSubFirstStationVehiclesPaths = [(i,v,p) for (i,v,p) in sFirstStationVehiclesPaths  if v == element[0] and p==element[1]]
    sSubNotFirstStationVehiclesPaths = [(i,v,p) for (i,v,p) in sNotFirstStationVehiclesPaths  if v == element[0] and p==element[1]]
    sSubNodesPotentialNodesOriginalVehiclesPaths = [(i,j,v,p) for (i,j,v,p) in sNodesPotentialNodesOriginalVehiclesPaths  if v == element[0] and p==element[1]]
    
    #solve subproblem
    d_subproblem_solution[element] = rfep_model.solve_rfep(
                            sNodesVehiclesPaths = sSubNodesVehiclesPaths,
                            sStationsVehiclesPaths = sSubStationsVehiclesPaths,
                            sOriginalStationsOwn = sOriginalStationsOwn,
                            sOriginalStationsPotential = sOriginalStationsPotential,
                            sSuppliers = sSuppliers,
                            sSuppliersRanges = sSuppliersRanges,
                            sOriginVehiclesPaths = sSubOriginVehiclesPaths,
                            sDestinationVehiclesPaths = sSubDestinationVehiclesPaths,
                            sSequenceNodesNodesVehiclesPaths = sSubSequenceNodesNodesVehiclesPaths,
                            sFirstStationVehiclesPaths = sSubFirstStationVehiclesPaths,
                            sNotFirstStationVehiclesPaths = sSubNotFirstStationVehiclesPaths,
                            sNodesPotentialNodesOriginalVehiclesPaths = sSubNodesPotentialNodesOriginalVehiclesPaths,
                            sOriginalStationsMirrorStations = sOriginalStationsMirrorStations,
                            sStationsSuppliers = sStationsSuppliers,
                            sSuppliersWithDiscount = sSuppliersWithDiscount,
                            sRanges = sRanges,
                            pStartInventory = pStartInventory,
                            pTargetInventory = pTargetInventory,
                            pSafetyStock = pSafetyStock,
                            pTankCapacity = pTankCapacity,
                            pMinRefuel = pMinRefuel,
                            pConsumptionMainRoute = pConsumptionMainRoute,
                            pConsumptionOOP = pConsumptionOOP,
                            pQuantityVehicles = pQuantityVehicles,
                            pPrice = pPrice,
                            pOpportunityCost = pOpportunityCost,
                            pVariableCost = pVariableCost,
                            pDistanceOOP = pDistanceOOP,
                            isONvInventory = True,
                            isONvRefuelQuantity = True,
                            isONvRefuel = True,
                            isONcInitialInventory = True,
                            isONcTargetInventory = True,
                            isONcMinInventory = True,
                            isONcLogicRefuel1 = True,
                            isONcLogicRefuel2 = True,
                            isONcMaxRefuel = True,
                            isONcInventoryBalance1 = True,
                            isONcInventoryBalance2 = True,
                            isONcInventoryBalance3 = True,
                            isONtotalRefuellingCost= True)                      

                         
    d_subproblem_status[element] = d_subproblem_solution[element][0]
    d_subproblem_ovInventory[element] = d_subproblem_solution[element][1]
    d_subproblem_ovRefuelQuantity[element] = d_subproblem_solution[element][2]
    d_subproblem_ovRefuel[element] = d_subproblem_solution[element][3]
    d_subproblem_oTotalRefuellingCost[element] = d_subproblem_solution[element][9]
    d_subproblem_oTotalCost[element] = d_subproblem_solution[element][12]
    
    d_subproblem_n_constraints[element] = d_subproblem_solution[element][13]
    d_subproblem_n_variables[element] = d_subproblem_solution[element][14]
    d_subproblem_n_integer_variables[element] = d_subproblem_solution[element][15]
    d_subproblem_n_binary_variables[element] = d_subproblem_solution[element][16]
    d_subproblem_model_fingerprint[element] = d_subproblem_solution[element][17]
    d_subproblem_model_runtime[element] = d_subproblem_solution[element][18]
    d_subproblem_model_MIPGap[element] = d_subproblem_solution[element][19]

    
    
    #%%
    #Export solution of the subproblem
    #Open Excel workbook
    scenario = "vhc - " + (element[0]) + " path -" + (str(element[1]))
    #this activates printing the output of each subproblem
    b_debug_subproblem = False
    if b_debug_subproblem:
        sheet_name = 'oNodesNodesVehiclesPaths'
        ws = workbook[sheet_name]
        index_row = ws.max_row
        
        #
        for (i,j,v,p) in sSubSequenceNodesNodesVehiclesPaths:
            if (j,p) in sStationsPaths:
                index_row=index_row +1
                ws.cell(row=index_row, column = 1, value=scenario)
                ws.cell(row=index_row, column = 2, value=i)
                ws.cell(row=index_row, column = 3, value=j)
                ws.cell(row=index_row, column = 4, value=v)
                ws.cell(row=index_row, column = 5, value=p)
                ws.cell(row=index_row, column = 6, value=d_subproblem_ovInventory[element][j,v,p])
                ws.cell(row=index_row, column = 7, value=d_subproblem_ovRefuelQuantity[element][j,v,p])
                ws.cell(row=index_row, column = 8, value=d_subproblem_ovRefuel[element][j,v,p])
                ws.cell(row=index_row, column = 9, value=pStartInventory[v,p])
                ws.cell(row=index_row, column = 10, value=pConsumptionRate[v])
                ws.cell(row=index_row, column = 11, value=pDistance[i,j,p])
                ws.cell(row=index_row, column = 12, value=pConsumptionMainRoute[i,j,v,p])
                ws.cell(row=index_row, column = 13, value=pDistanceOOP[j,p])
                ws.cell(row=index_row, column = 14, value=pConsumptionOOP[j,v,p])
                ws.cell(row=index_row, column = 15, value=pPrice[j])
                ws.cell(row=index_row, column = 16, value=0) #this is a placeholder in case I need this to run single corridor
                ws.cell(row=index_row, column = 17, value=pQuantityVehicles[v,p])
                ws.cell(row=index_row, column = 18, value=pVariableCost[v])
                ws.cell(row=index_row, column = 19, value=pOpportunityCost[v])
        
        #Print total cost
        sheet_name = 'oTotalCost'
        ws = workbook[sheet_name]
        index_row = ws.max_row + 1
        ws.cell(row = index_row, column = 1, value = scenario)
        ws.cell(row = index_row, column = 2, value = d_subproblem_oTotalRefuellingCost[element])
        ws.cell(row = index_row, column = 3, value = 0)
        ws.cell(row = index_row, column = 4, value = 0)
        ws.cell(row = index_row, column = 5, value = d_subproblem_oTotalCost[element])
    
    b_print_stats=True
    if b_print_stats:    
        #print solution statistics
        sheet_name = 'oScenarioStats'
        ws = workbook[sheet_name]
        index_row = ws.max_row + 1
        ws.cell(row = index_row, column = 1, value = file)
        ws.cell(row = index_row, column = 2, value = scenario)
        ws.cell(row = index_row, column = 12, value = d_subproblem_model_runtime[element])
        ws.cell(row = index_row, column = 13, value = d_subproblem_n_constraints[element])
        ws.cell(row = index_row, column = 14, value = d_subproblem_n_variables[element])
        ws.cell(row = index_row, column = 15, value = d_subproblem_n_integer_variables[element])
        ws.cell(row = index_row, column = 16, value = d_subproblem_n_binary_variables[element])
        ws.cell(row = index_row, column = 17, value = d_subproblem_model_fingerprint[element])
        ws.cell(row = index_row, column = 18, value = d_subproblem_model_MIPGap[element])
       
workbook.save(output_file)
       
    
                                            

