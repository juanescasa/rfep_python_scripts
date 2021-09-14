# -*- coding: utf-8 -*-
"""
Created on Wed Apr 28 11:03:54 2021

@author: calle
"""
#import all what you need
#%%
import pandas as pd
import gurobipy as gp
from gurobipy import GRB
import os
import time
import openpyxl
import sys
from read_data_rfep import *

#%%




l_time_track.append(('start_model_generation', time.time()))

#Define the model
m = gp.Model()

#Define variables
vInventory = m.addVars(sNodesVehiclesPaths, name = 'vInventory')
vRefuelQuantity = m.addVars(sStationsVehiclesPaths, name = 'vRefuelQuantity')
vRefuel = m.addVars(sStationsVehiclesPaths, vtype = GRB.BINARY, name = 'vRefuel')
vQuantityUnitsCapacity = m.addVars(sOriginalStationsOwn, vtype = GRB.INTEGER, name = 'vQuantityUnitsCapacity')
vLocate = m.addVars(sOriginalStationsPotential, vtype = GRB.BINARY, name = 'vLocate')
vQuantityPurchased = m.addVars(sSuppliers, name = 'vQuantityPurchased')
vQuantityPurchasedRange = m.addVars(sSuppliersRanges, name = 'vQuantityPurchasedRange')
vPurchasedRange = m.addVars(sSuppliersRanges, vtype = GRB.BINARY, name='vPurchasedRange')

#Define Constraints
#Refuelling Logic Constraints
cInitialInventory = m.addConstrs((vInventory[i,v,p]==pStartInventory[v,p] 
                                  for (i,v,p) in sOriginVehiclesPaths), name = 'initialInventory')
cTargetInventory = m.addConstrs((vInventory[i,v,p]==pTargetInventory[v,p] 
                                  for (i,v,p) in sDestinationVehiclesPaths), name = 'targetInventory')
cMinInventory = m.addConstrs((vInventory[i,v,p]>=pSafetyStock[v] 
                               for (i,v,p) in sStationsVehiclesPaths), name ='minInventory')
cLogicRefuel1 = m.addConstrs((vRefuelQuantity[i,v,p]<=(pTankCapacity[v]-pSafetyStock[v])*vRefuel[i,v,p] 
                               for (i,v,p) in sStationsVehiclesPaths), name = 'logicRefuel1')
cLogicRefuel2 = m.addConstrs((vRefuelQuantity[i,v,p]>=pMinRefuel[v]*vRefuel[i,v,p]
                               for (i,v,p) in sStationsVehiclesPaths), name = 'logicRefuel2')
cMaxRefuel = m.addConstrs((vInventory[i,v,p] + vRefuelQuantity[i,v,p] <= pTankCapacity[v] 
                            for (i,v,p) in sStationsVehiclesPaths), name = 'maxRefuel')
cInventoryBalance1 = m.addConstrs((vInventory[j,v,p] == pStartInventory[v,p] - (pConsumptionMainRoute[i,j,v,p] + pConsumptionOOP[j,v,p]*vRefuel[j,v,p]) 
                                    for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if (j,v,p) in sFirstStationVehiclesPaths), name = 'inventoryBalance1')
cInventoryBalance2 = m.addConstrs((vInventory[j,v,p] == vInventory[i,v,p] + vRefuelQuantity[i,v,p] - 
               (pConsumptionOOP[i,v,p]*vRefuel[i,v,p] + pConsumptionMainRoute[i,j,v,p] + pConsumptionOOP[j,v,p]*vRefuel[j,v,p]) 
               for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if (j,v,p) in sNotFirstStationVehiclesPaths), name = 'inventoryBalance2')
cInventoryBalance3 = m.addConstrs((vInventory[j,v,p]==vInventory[i,v,p]+vRefuelQuantity[i,v,p] - (pConsumptionMainRoute[i,j,v,p] + pConsumptionOOP[i,v,p]*vRefuel[i,v,p]) 
                                   for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if (j,v,p) in sDestinationVehiclesPaths), name = 'inventoryBalance3')
# #Location Logic Constraints
cLogicLocation = m.addConstrs((vRefuel[j,v,p] <= vLocate[i] for (j,i,v,p) in sNodesPotentialNodesOriginalVehiclesPaths), name = 'logicLocation')
# #Valid Inequality
cLogicLocation2 = m.addConstrs((vRefuelQuantity[j,v,p] <= (pTankCapacity[v]-pSafetyStock[v])*vLocate[i] for (j,i,v,p) in sNodesPotentialNodesOriginalVehiclesPaths), name = 'logicLocation2')

cStationCapacity = m.addConstrs(((gp.quicksum(pQuantityVehicles[v,p] * vRefuelQuantity[j,v,p] for (j,v,p) in sStationsVehiclesPaths if (i,j) in sOriginalStationsMirrorStations))
                                  <=pStationCapacity[i] + pStationUnitCapacity[i]*vQuantityUnitsCapacity[i]  for i in sOriginalStationsOwn), name = 'stationCapacity')
# #Supplier Logic constraints

cQuantityPurchased = m.addConstrs((vQuantityPurchased[l] == gp.quicksum(pQuantityVehicles[v,p]*vRefuelQuantity[i,v,p] for (i,v,p) in sStationsVehiclesPaths if (i,l) in sStationsSuppliers) for l in sSuppliers), name = 'quantityPurchased')

cMinimumQuantitySupplier = m.addConstrs((vQuantityPurchased[l]>=pMinimumPurchaseQuantity[l] for l in sSuppliers), name = 'minimumQuantitySupplier')

cMinQuantityRange = m.addConstrs((vQuantityPurchasedRange[l,g]>= vPurchasedRange[l,g]*pLowerQuantityDiscount[l,g] for (l,g) in sSuppliersRanges), name = 'minQuantityRange')

cMaxQuantityRange = m.addConstrs((vQuantityPurchasedRange[l,g]<= vPurchasedRange[l,g]*pUpperQuantityDiscount[l,g] for (l,g) in sSuppliersRanges), name = 'maxQuantityRange')

cUniqueQuantityRange = m.addConstrs((gp.quicksum(vQuantityPurchasedRange[l,g] for g in sRanges if (l,g) in sSuppliersRanges) == vQuantityPurchased[l] for l in sSuppliersWithDiscount), name = 'uniqueQuantityRange')

cUniqueRange = m.addConstrs((gp.quicksum(vPurchasedRange[l,g] for g in sRanges if (l, g) in sSuppliersRanges) == 1 for l in sSuppliersWithDiscount), name = 'uniqueRange')

#Objective Function
totalRefuellingCost = gp.quicksum(pQuantityVehicles[v,p]*(vRefuelQuantity[i,v,p]*pPrice[i] + (pOpportunityCost[v] + 2*pVariableCost[v]*pDistanceOOP[i,p])*vRefuel[i,v,p])  
                                   for (i,v,p) in sStationsVehiclesPaths)
totalLocationCost = gp.quicksum(pLocationCost[i]*vLocate[i] for i in sOriginalStationsPotential) + gp.quicksum(pCostUnitCapacity[i]*vQuantityUnitsCapacity[i] for i in sOriginalStationsOwn)
totalDiscount = gp.quicksum(pDiscount[l,g]*vQuantityPurchasedRange[l,g] for (l,g) in sSuppliersRanges)

m.setObjective(totalRefuellingCost + totalLocationCost - totalDiscount, GRB.MINIMIZE )
#m.setObjective(0, GRB.MINIMIZE )
#This allows to get a definitive conclusion between unbounded or infeasible model.
m.Params.DualReductions = 0
l_time_track.append(('start_optimization', time.time()))
m.optimize()
l_time_track.append(('start_export_output', time.time()))
#Dealing with infeasability
status = m.status
if status == GRB.INFEASIBLE:
    print("The model is infeasible")
    m.computeIIS()
    for c in m.getConstrs():
        if c.IISConstr:
            print('%s' % c.constrName)
    #this stop the execution of this file.
    sys.exit()
    
#dsp.Params.LogToConsole = 0

#print solution
#save variables values in output dictionaries
#%%
scenario ='standard'

ovInventory = m.getAttr('x', vInventory)
ovRefuelQuantity = m.getAttr('x', vRefuelQuantity)
ovRefuel = m.getAttr('x', vRefuel)
ovQuantityUnitsCapacity = m.getAttr('x', vQuantityUnitsCapacity)
ovLocate = m.getAttr('x', vLocate)
ovQuantityPurchased = m.getAttr('x', vQuantityPurchased)
ovQuantityPurchasedRange = m.getAttr('x', vQuantityPurchasedRange)
ovPurchasedRange = m.getAttr('x', vPurchasedRange)

output_file = os.path.join("..", "output", "outputRFEP.xlsx")
sheet = 'oNodesNodesVehiclesPaths'
workbook = openpyxl.load_workbook(output_file)
sh_remove = workbook[sheet]
workbook.remove(sh_remove)
ws = workbook.create_sheet(sheet)

def write_column_titles(worksheet, list_names):
    index_column=0
    for index in list_names:
        index_column=index_column+1
        worksheet.cell(row=1, column=index_column, value = index)

oColumnsNodesNodesVehiclesPaths = ["Scenario", "COD_NODE1", "COD_NODE2", "COD_VEHICLE", "COD_PATH", "vInventory", 
                                    "vRefuelQuantity", "vRefuel", "pStartInventory", "pConsumptionRate", "pDistance", 
                                    "pConsumptionMainRoute", "pDistanceOOP", "pConsumptionOOP", "pPrice", "pRefuelQuantity", 
                                    "pQuantityVehicles", "pVariableCost", "pOpportunityCost"]

write_column_titles(ws, oColumnsNodesNodesVehiclesPaths)
index_row=1
for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths:
    if (j,p) in sStationsPaths:
        index_row=index_row +1
        ws.cell(row=index_row, column = 1, value=scenario)
        ws.cell(row=index_row, column = 2, value=i)
        ws.cell(row=index_row, column = 3, value=j)
        ws.cell(row=index_row, column = 4, value=v)
        ws.cell(row=index_row, column = 5, value=p)
        ws.cell(row=index_row, column = 6, value=ovInventory[j,v,p])
        ws.cell(row=index_row, column = 7, value=ovRefuelQuantity[j,v,p])
        ws.cell(row=index_row, column = 8, value=ovRefuel[j,v,p])
        ws.cell(row=index_row, column = 9, value=pStartInventory[v,p])
        ws.cell(row=index_row, column = 10, value=pConsumptionRate[v])
        ws.cell(row=index_row, column = 11, value=pDistance[i,j,p])
        ws.cell(row=index_row, column = 12, value=pConsumptionMainRoute[i,j,v,p])
        ws.cell(row=index_row, column = 13, value=pDistanceOOP[j,p])
        ws.cell(row=index_row, column = 14, value=pConsumptionOOP[j,v,p])
        ws.cell(row=index_row, column = 15, value=pPrice[j])
        ws.cell(row=index_row, column = 16, value=0) #this is a placeholder in case I need this to run single corridor
        ws.cell(row=index_row, column = 17, value=pQuantityVehicles[v,p])
        ws.cell(row=index_row, column = 18, value=pVariableCost[v])
        ws.cell(row=index_row, column = 19, value=pOpportunityCost[v])

#Export the location output table

sheet = 'oOriginalStationsOwn'
sh_remove = workbook[sheet]
workbook.remove(sh_remove)
ws = workbook.create_sheet(sheet)

oColumnsOriginalStationsOwn = ["Scenario", "COD_NODE", "vLocate", "vQuantityUnitsCapacity", "pStationCapacity", "pStationUnitCapacity", "pLocationCost", "pCostUnitCapacity"]
write_column_titles(ws, oColumnsOriginalStationsOwn)
index_row=1

for i in sOriginalStationsOwn:
    index_row = index_row+1
    ws.cell(row=index_row, column = 1, value = scenario)
    ws.cell(row=index_row, column = 2, value = i)
    if i in sOriginalStationsPotential:
        ws.cell(row=index_row, column = 3, value = ovLocate[i])
        ws.cell(row=index_row, column = 7, value = pLocationCost[i])
    else:
        ws.cell(row=index_row, column = 3, value = 0)
        ws.cell(row=index_row, column = 7, value = 0)
    ws.cell(row=index_row, column = 4, value = ovQuantityUnitsCapacity[i])
    ws.cell(row=index_row, column = 5, value = pStationCapacity[i])
    ws.cell(row=index_row, column = 6, value = pStationUnitCapacity[i])
    ws.cell(row=index_row, column = 8, value = pCostUnitCapacity[i])

oTotalRefuellingCost = sum(pQuantityVehicles[v,p]*(ovRefuelQuantity[i,v,p]*pPrice[i] + (pOpportunityCost[v] + 2*pVariableCost[v]*pDistanceOOP[i,p])*ovRefuel[i,v,p])  
                                   for (i,v,p) in sStationsVehiclesPaths)
oTotalLocationCost = sum(pLocationCost[i]*ovLocate[i] for i in sOriginalStationsPotential) + sum(pCostUnitCapacity[i]*ovQuantityUnitsCapacity[i] for i in sOriginalStationsOwn)
oTotalDiscount = sum(pDiscount[l,g]*ovQuantityPurchasedRange[l,g] for (l,g) in sSuppliersRanges)
oTotalCost = oTotalRefuellingCost + oTotalLocationCost - oTotalDiscount

sheet = 'oTotalCost'
sh_remove = workbook[sheet]
workbook.remove(sh_remove)
ws = workbook.create_sheet(sheet)

oColumnsTotalCost = ["Scenario", "TotalRefuellingCost", "TotalLocationCost", "TotalDiscount", "TotalCost"]
write_column_titles(ws, oColumnsTotalCost)

ws.cell(row=2, column=1, value = scenario)
ws.cell(row=2, column=2, value = oTotalRefuellingCost)
ws.cell(row=2, column=3, value = oTotalLocationCost)
ws.cell(row=2, column=4, value = oTotalDiscount)
ws.cell(row=2, column=5, value = oTotalCost)



l_time_track.append(('end', time.time()))

#Export scenario characteristics
l_process = ['read_data', 'calculate_parameters', 'generate_model', 'optimize_model', 'export_output' ]
l_process_time = []
for row in range(len(l_process)):
    l_process_time.append((l_process[row], (l_time_track[row+1][1]-l_time_track[row][1])))

n_vehicles=len(df_vehicles_paths['COD_VEHICLE'].unique())
n_paths=len(df_vehicles_paths['COD_PATH'].unique())
n_avg_stations_path = df_nodes_paths.groupby('COD_PATH')['COD_NODE1'].count().subtract(2).mean()
n_candidate_locations = len(sOriginalStationsPotential)

n_constraints = m.NumConstrs
n_variables = m.Numvars
n_integer_variables = m.NumIntVars
n_binary_variables = m.NumBinVars
model_fingerprint = m.Fingerprint
model_runtime = m.Runtime
model_MIPGap = m.Mipgap

stats_to_export = [file, n_vehicles, n_paths, n_avg_stations_path, n_candidate_locations,
                   l_process_time[0][1], l_process_time[1][1], l_process_time[2][1],
                   l_process_time[3][1], l_process_time[4][1], model_runtime, 
                   n_constraints, n_variables, n_integer_variables, n_binary_variables, 
                   model_MIPGap, model_fingerprint]

sheet = 'oScenarioStats'
ws = workbook[sheet]

#define which row to write the data
row_to_write = ws.max_row + 1
for index in range(len(stats_to_export)):
    ws.cell(row=row_to_write, column = index+1, value= stats_to_export[index])

workbook.save(output_file)

#m.write("RFEP.lp")
#m.write("RFEP.sol")
