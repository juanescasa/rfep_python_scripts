# -*- coding: utf-8 -*-
"""
Created on Wed Apr 28 11:03:54 2021

@author: calle
"""
#import all what you need
#%%
import pandas as pd
import gurobipy as gp
from gurobipy import GRB
import os
import time
import openpyxl
import sys

start_time = time.time()

#Read tables as dataframes
#file = os.path.join("..", "data", "Data Model Generated Network-13.xlsm")
file = "C:\OneDrive - Deakin University\OD\calle test\Disun Applications\Gurobi Applications\data\Data Model Generated Network-16.xlsm"
sh = 'MaeVehicles'
df_vehicles = pd.read_excel(file, sheet_name = sh)

#Read inputs of mathematical model
##Define the elements of the mathematical model (sets and parameters)

sVehicles=[]
pSafetyStock = {}
pTankCapacity = {}
pConsumptionRate = {}
pMinRefuel = {}
pVariableCost = {}
pOpportunityCost = {}

##Assign the values from dataframes to elements of mathematical model
for index in range(df_vehicles.shape[0]):
    v = df_vehicles['COD_VEHICLE'][index]
    sVehicles.append(v)
    pSafetyStock[v]=df_vehicles['pSafetyStock'][index]
    pTankCapacity[v]=df_vehicles['pTankCapacity'][index]
    pConsumptionRate[v]=df_vehicles['pConsumptionRate'][index]
    pMinRefuel[v]=df_vehicles['pMinRefuel'][index]
    pVariableCost[v]=df_vehicles['pVariableCost'][index]
    pOpportunityCost[v]=df_vehicles['pOpportunityCost'][index]

sh = 'MaeSuppliers'
df_suppliers = pd.read_excel(file, sheet_name = sh)

sSuppliers=[]
sSuppliersWithDiscount = []
pMinimumPurchaseQuantity = {}

for index in range(df_suppliers.shape[0]):
    l = df_suppliers['COD_SUPPLIER'][index]
    sSuppliers.append(l)
    if df_suppliers['pSuppliersWithDiscount'][index] ==1:
        sSuppliersWithDiscount.append(l)
    pMinimumPurchaseQuantity[l]=df_suppliers['pMinimumPurchaseQuantity'][index]

sh = 'MaeRanges'
df_ranges = pd.read_excel(file, sheet_name = sh)

sRanges=[]


for index in range(df_ranges.shape[0]):
    l = df_ranges['COD_RANGE'][index]
    sRanges.append(l)


sh = 'NodesNodes'
df_nodes_nodes = pd.read_excel(file, sheet_name = sh)


sOriginalStationsMirrorStations = []

for index in range(df_nodes_nodes.shape[0]):
    i = df_nodes_nodes['COD_NODE1'][index]
    j = df_nodes_nodes['COD_NODE2'][index]
    sOriginalStationsMirrorStations.append((i,j))


sh = 'SubStations'
df_substations = pd.read_excel(file, sheet_name = sh)

sOriginalStationsOwn = []
sOriginalStationsPotential = []
sAuxStationsSuppliers = []
sAuxStations = []

pStationCapacity = {}
pStationUnitCapacity = {}
pCostUnitCapacity = {}
pLocationCost = {}
pAuxPrice = {}

for index in range(df_substations.shape[0]):
    i = df_substations['COD_NODE'][index]
    l = df_substations['COD_SUPPLIER'][index]
    if df_substations['COD_SUPPLIER'][index] == 'OWN':
        sOriginalStationsOwn.append(i)
        pStationCapacity[i]= df_substations['pStationCapacity'][index]
        pStationUnitCapacity[i] = df_substations['pStationUnitCapacity'][index]
        pCostUnitCapacity[i] = df_substations['pCostUnitCapacity'][index]
        if df_substations['isPotential'][index] == 1:
            sOriginalStationsPotential.append(i)
            pLocationCost[i] = df_substations['pLocationCost'][index]
    
    sAuxStations.append(i)
    sAuxStationsSuppliers.append((i,l))
    pAuxPrice[i]=df_substations['pPriceOriginal'][index]

sh = 'VehiclesPaths'
df_vehicles_paths = pd.read_excel(file, sheet_name = sh)
 
sVehiclesPaths = []

pAuxStartInventory = {}
pTargetInventory = {}
pAuxQuantityVehicles = {}

for index in range(df_vehicles_paths.shape[0]):
    v = df_vehicles_paths['COD_VEHICLE'][index]
    p = df_vehicles_paths['COD_PATH'][index]
    sVehiclesPaths.append((v,p))
    pAuxStartInventory[v,p] = df_vehicles_paths['Start Inventory'][index]
    pTargetInventory[v,p] = df_vehicles_paths['Target Inventory'][index]
    pAuxQuantityVehicles[v,p] = df_vehicles_paths['pQuantityVehicles'][index]
    

sh = 'NodesPaths'
df_nodes_paths = pd.read_excel(file, sheet_name = sh)

sAuxNodesPaths = []
sAuxOriginPaths = []
sAuxDestinationPaths = []
sAuxFirstStationPaths = []

pDistanceOOP = {}

for index in range(df_nodes_paths.shape[0]):
    i = df_nodes_paths['COD_NODE1'][index]
    p = df_nodes_paths['COD_PATH'][index]
    sAuxNodesPaths.append((i,p))
    if df_nodes_paths['pOriginPath'][index]==1:
        sAuxOriginPaths.append((i,p))
    if df_nodes_paths['pDestinationPath'][index]==1:
        sAuxDestinationPaths.append((i,p))
    if df_nodes_paths['pFirstStation'][index]==1:
        sAuxFirstStationPaths.append((i,p))
    if df_nodes_paths['pOriginPath'][index]==0 and df_nodes_paths['pDestinationPath'][index]==0:
        pDistanceOOP[i,p]=df_nodes_paths['pDistanceOOP'][index]

sh = 'SuppliersRanges'
df_suppliers_ranges = pd.read_excel(file, sheet_name = sh)

sSuppliersRanges = []

pLowerQuantityDiscount = {}
pUpperQuantityDiscount = {}
pDiscount = {}

for index in range(df_suppliers_ranges.shape[0]):
    l = df_suppliers_ranges['COD_SUPPLIER'][index]
    g = df_suppliers_ranges['COD_RANGE'][index]
    sSuppliersRanges.append((l,g))
    pLowerQuantityDiscount[l,g] = df_suppliers_ranges['pLowerQuantityDiscount'][index]
    pUpperQuantityDiscount[l,g] = df_suppliers_ranges['pUpperQuantityDiscount'][index]
    pDiscount[l,g] = df_suppliers_ranges['pDiscount'][index]

sh = 'NodesNodesPaths'
df_nodes_nodes_paths = pd.read_excel(file, sheet_name = sh)

sAuxNodesNodesPaths = []
pDistance = {}

for index in range(df_nodes_nodes_paths.shape[0]):
    i = df_nodes_nodes_paths['COD_NODE1'][index]
    j = df_nodes_nodes_paths['COD_NODE2'][index]
    p = df_nodes_nodes_paths['COD_PATH'][index]
    sAuxNodesNodesPaths.append((i,j,p))
    pDistance[i,j,p] = df_nodes_nodes_paths['pDistance'][index]

#%%
#calculated sets
sStations = [i for (ii,i) in sOriginalStationsMirrorStations]
sStationsPaths = [(i,p) for (i,p) in sAuxNodesPaths if i in sStations]
sNodesVehiclesPaths = [(i,v,p) for (i,p) in sAuxNodesPaths for (v,pp) in sVehiclesPaths  if p == pp]
sStationsVehiclesPaths = [(i,v,p) for (i,p) in sAuxNodesPaths for (v,pp) in sVehiclesPaths  if p == pp if i in sStations]
sDestinationVehiclesPaths = [(i,v,p) for (i,p) in sAuxDestinationPaths for(v,pp) in sVehiclesPaths  if p == pp]
sOriginVehiclesPaths = [(i,v,p) for (i,p) in sAuxOriginPaths for(v,pp) in sVehiclesPaths  if p == pp]
sFirstStationVehiclesPaths = [(i,v,p) for (i,p) in sAuxFirstStationPaths for(v,pp) in sVehiclesPaths  if p == pp]
sAuxNotFirstStationVehiclesPaths = ((set(sNodesVehiclesPaths)-set(sFirstStationVehiclesPaths))-set(sOriginVehiclesPaths))-set(sDestinationVehiclesPaths)
sNotFirstStationVehiclesPaths =  list(sAuxNotFirstStationVehiclesPaths) 
sNodesPotentialNodesOriginalVehiclesPaths = [(j,i,v,p) for (i,j) in sOriginalStationsMirrorStations for (jj,v,p) in sNodesVehiclesPaths if j==jj if i in sOriginalStationsPotential]
sSequenceNodesNodesVehiclesPaths = [(i,j,v,p) for (i,j,p) in sAuxNodesNodesPaths for (v,pp) in sVehiclesPaths if p == pp]
sStationsSuppliers = [(j,l) for (i,j) in sOriginalStationsMirrorStations for (ii,l) in sAuxStationsSuppliers if i == ii]

#Auxiliar parameters
pFlowFactor = 1
pPercentageStartInventory = 1

#calculated parameters
#This is called dictionary comprehension: https://www.netguru.com/codestories/python-list-comprehension-dictionary
pPrice={i: pAuxPrice[ii] for (ii,i) in sOriginalStationsMirrorStations}
pConsumptionOOP = {(i,v,p): pDistanceOOP[i,p]*pConsumptionRate[v] for (i,p) in sStationsPaths for (v,pp) in sVehiclesPaths if p == pp}
pConsumptionMainRoute = {(i,j,v,p): pDistance[i,j,p] * pConsumptionRate[v] for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths}
pQuantityVehicles = {(v,p): pFlowFactor*pAuxQuantityVehicles[v,p] for (v,p) in sVehiclesPaths}
pStartInventory =  {(v,p): pPercentageStartInventory*pAuxStartInventory[v,p] for (v,p) in sVehiclesPaths}

#Define the model
m = gp.Model()

#Define variables
vInventory = m.addVars(sNodesVehiclesPaths, name = 'vInventory')
vRefuelQuantity = m.addVars(sStationsVehiclesPaths, name = 'vRefuelQuantity')
vRefuel = m.addVars(sStationsVehiclesPaths, vtype = GRB.BINARY, name = 'vRefuel')
vQuantityUnitsCapacity = m.addVars(sOriginalStationsOwn, vtype = GRB.INTEGER, name = 'vQuantityUnitsCapacity')
vLocate = m.addVars(sOriginalStationsPotential, vtype = GRB.BINARY, name = 'vLocate')
vQuantityPurchased = m.addVars(sSuppliers, name = 'vQuantityPurchased')
vQuantityPurchasedRange = m.addVars(sSuppliersRanges, name = 'vQuantityPurchasedRange')
vPurchasedRange = m.addVars(sSuppliersRanges, vtype = GRB.BINARY, name='vPurchasedRange')

#Define Constraints
#Refuelling Logic Constraints
cInitialInventory = m.addConstrs((vInventory[i,v,p]==pStartInventory[v,p] 
                                  for (i,v,p) in sOriginVehiclesPaths), name = 'initialInventory')
cTargetInventory = m.addConstrs((vInventory[i,v,p]==pTargetInventory[v,p] 
                                  for (i,v,p) in sDestinationVehiclesPaths), name = 'targetInventory')
cMinInventory = m.addConstrs((vInventory[i,v,p]>=pSafetyStock[v] 
                               for (i,v,p) in sStationsVehiclesPaths), name ='minInventory')
cLogicRefuel1 = m.addConstrs((vRefuelQuantity[i,v,p]<=(pTankCapacity[v]-pSafetyStock[v])*vRefuel[i,v,p] 
                               for (i,v,p) in sStationsVehiclesPaths), name = 'logicRefuel1')
cLogicRefuel2 = m.addConstrs((vRefuelQuantity[i,v,p]>=pMinRefuel[v]*vRefuel[i,v,p]
                               for (i,v,p) in sStationsVehiclesPaths), name = 'logicRefuel2')
cMaxRefuel = m.addConstrs((vInventory[i,v,p] + vRefuelQuantity[i,v,p] <= pTankCapacity[v] 
                            for (i,v,p) in sStationsVehiclesPaths), name = 'maxRefuel')
cInventoryBalance1 = m.addConstrs((vInventory[j,v,p] == pStartInventory[v,p] - (pConsumptionMainRoute[i,j,v,p] + pConsumptionOOP[j,v,p]*vRefuel[j,v,p]) 
                                    for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if (j,v,p) in sFirstStationVehiclesPaths), name = 'inventoryBalance1')
cInventoryBalance2 = m.addConstrs((vInventory[j,v,p] == vInventory[i,v,p] + vRefuelQuantity[i,v,p] - 
               (pConsumptionOOP[i,v,p]*vRefuel[i,v,p] + pConsumptionMainRoute[i,j,v,p] + pConsumptionOOP[j,v,p]*vRefuel[j,v,p]) 
               for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if (j,v,p) in sNotFirstStationVehiclesPaths), name = 'inventoryBalance2')
cInventoryBalance3 = m.addConstrs((vInventory[j,v,p]==vInventory[i,v,p]+vRefuelQuantity[i,v,p] - (pConsumptionMainRoute[i,j,v,p] + pConsumptionOOP[i,v,p]*vRefuel[i,v,p]) 
                                   for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths if (j,v,p) in sDestinationVehiclesPaths), name = 'inventoryBalance3')
# #Location Logic Constraints
cLogicLocation = m.addConstrs((vRefuel[j,v,p] <= vLocate[i] for (j,i,v,p) in sNodesPotentialNodesOriginalVehiclesPaths), name = 'logicLocation')
# #Valid Inequality
cLogicLocation2 = m.addConstrs((vRefuelQuantity[j,v,p] <= (pTankCapacity[v]-pSafetyStock[v])*vLocate[i] for (j,i,v,p) in sNodesPotentialNodesOriginalVehiclesPaths), name = 'logicLocation2')

cStationCapacity = m.addConstrs(((gp.quicksum(pQuantityVehicles[v,p] * vRefuelQuantity[j,v,p] for (j,v,p) in sStationsVehiclesPaths if (i,j) in sOriginalStationsMirrorStations))
                                  <=pStationCapacity[i] + pStationUnitCapacity[i]*vQuantityUnitsCapacity[i]  for i in sOriginalStationsOwn), name = 'stationCapacity')
# #Supplier Logic constraints

cQuantityPurchased = m.addConstrs((vQuantityPurchased[l] == gp.quicksum(pQuantityVehicles[v,p]*vRefuelQuantity[i,v,p] for (i,v,p) in sStationsVehiclesPaths if (i,l) in sStationsSuppliers) for l in sSuppliers), name = 'quantityPurchased')

cMinimumQuantitySupplier = m.addConstrs((vQuantityPurchased[l]>=pMinimumPurchaseQuantity[l] for l in sSuppliers), name = 'minimumQuantitySupplier')

cMinQuantityRange = m.addConstrs((vQuantityPurchasedRange[l,g]>= vPurchasedRange[l,g]*pLowerQuantityDiscount[l,g] for (l,g) in sSuppliersRanges), name = 'minQuantityRange')

cMaxQuantityRange = m.addConstrs((vQuantityPurchasedRange[l,g]<= vPurchasedRange[l,g]*pUpperQuantityDiscount[l,g] for (l,g) in sSuppliersRanges), name = 'maxQuantityRange')

cUniqueQuantityRange = m.addConstrs((gp.quicksum(vQuantityPurchasedRange[l,g] for g in sRanges if (l,g) in sSuppliersRanges) == vQuantityPurchased[l] for l in sSuppliersWithDiscount), name = 'uniqueQuantityRange')

cUniqueRange = m.addConstrs((gp.quicksum(vPurchasedRange[l,g] for g in sRanges if (l, g) in sSuppliersRanges) == 1 for l in sSuppliersWithDiscount), name = 'uniqueRange')

#Objective Function
totalRefuellingCost = gp.quicksum(pQuantityVehicles[v,p]*(vRefuelQuantity[i,v,p]*pPrice[i] + (pOpportunityCost[v] + 2*pVariableCost[v]*pDistanceOOP[i,p])*vRefuel[i,v,p])  
                                   for (i,v,p) in sStationsVehiclesPaths)
totalLocationCost = gp.quicksum(pLocationCost[i]*vLocate[i] for i in sOriginalStationsPotential) + gp.quicksum(pCostUnitCapacity[i]*vQuantityUnitsCapacity[i] for i in sOriginalStationsOwn)
totalDiscount = gp.quicksum(pDiscount[l,g]*vQuantityPurchasedRange[l,g] for (l,g) in sSuppliersRanges)

m.setObjective(totalRefuellingCost + totalLocationCost - totalDiscount, GRB.MINIMIZE )
#m.setObjective(0, GRB.MINIMIZE )
#This allows to get a definitive conclusion between unbounded or infeasible model.
m.Params.DualReductions = 0

m.optimize()

#Dealing with infeasability
status = m.status
if status == GRB.INFEASIBLE:
    print("The model is infeasible")
    m.computeIIS()
    for c in m.getConstrs():
        if c.IISConstr:
            print('%s' % c.constrName)
    #this stop the execution of this file.
    sys.exit()
    
#dsp.Params.LogToConsole = 0

#print solution
#save variables values in output dictionaries
scenario ='standard'

ovInventory = m.getAttr('x', vInventory)
ovRefuelQuantity = m.getAttr('x', vRefuelQuantity)
ovRefuel = m.getAttr('x', vRefuel)
ovQuantityUnitsCapacity = m.getAttr('x', vQuantityUnitsCapacity)
ovLocate = m.getAttr('x', vLocate)
ovQuantityPurchased = m.getAttr('x', vQuantityPurchased)
ovQuantityPurchasedRange = m.getAttr('x', vQuantityPurchasedRange)
ovPurchasedRange = m.getAttr('x', vPurchasedRange)

output_file = os.path.join("..", "output", "outputRFEP.xlsx")
sheet = 'oNodesNodesVehiclesPaths'
workbook = openpyxl.load_workbook(output_file)
sh_remove = workbook[sheet]
workbook.remove(sh_remove)
ws = workbook.create_sheet(sheet)

def write_column_titles(worksheet, list_names):
    index_column=0
    for index in list_names:
        index_column=index_column+1
        worksheet.cell(row=1, column=index_column, value = index)

oColumnsNodesNodesVehiclesPaths = ["Scenario", "COD_NODE1", "COD_NODE2", "COD_VEHICLE", "COD_PATH", "vInventory", 
                                    "vRefuelQuantity", "vRefuel", "pStartInventory", "pConsumptionRate", "pDistance", 
                                    "pConsumptionMainRoute", "pDistanceOOP", "pConsumptionOOP", "pPrice", "pRefuelQuantity", 
                                    "pQuantityVehicles", "pVariableCost", "pOpportunityCost"]

write_column_titles(ws, oColumnsNodesNodesVehiclesPaths)
index_row=1
for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths:
    if (j,p) in sStationsPaths:
        index_row=index_row +1
        ws.cell(row=index_row, column = 1, value=scenario)
        ws.cell(row=index_row, column = 2, value=i)
        ws.cell(row=index_row, column = 3, value=j)
        ws.cell(row=index_row, column = 4, value=v)
        ws.cell(row=index_row, column = 5, value=p)
        ws.cell(row=index_row, column = 6, value=ovInventory[j,v,p])
        ws.cell(row=index_row, column = 7, value=ovRefuelQuantity[j,v,p])
        ws.cell(row=index_row, column = 8, value=ovRefuel[j,v,p])
        ws.cell(row=index_row, column = 9, value=pStartInventory[v,p])
        ws.cell(row=index_row, column = 10, value=pConsumptionRate[v])
        ws.cell(row=index_row, column = 11, value=pDistance[i,j,p])
        ws.cell(row=index_row, column = 12, value=pConsumptionMainRoute[i,j,v,p])
        ws.cell(row=index_row, column = 13, value=pDistanceOOP[j,p])
        ws.cell(row=index_row, column = 14, value=pConsumptionOOP[j,v,p])
        ws.cell(row=index_row, column = 15, value=pPrice[j])
        ws.cell(row=index_row, column = 16, value=0) #this is a placeholder in case I need this to run single corridor
        ws.cell(row=index_row, column = 17, value=pQuantityVehicles[v,p])
        ws.cell(row=index_row, column = 18, value=pVariableCost[v])
        ws.cell(row=index_row, column = 19, value=pOpportunityCost[v])

#Export the location output table

sheet = 'oOriginalStationsOwn'
sh_remove = workbook[sheet]
workbook.remove(sh_remove)
ws = workbook.create_sheet(sheet)

oColumnsOriginalStationsOwn = ["Scenario", "COD_NODE", "vLocate", "vQuantityUnitsCapacity", "pStationCapacity", "pStationUnitCapacity", "pLocationCost", "pCostUnitCapacity"]
write_column_titles(ws, oColumnsOriginalStationsOwn)
index_row=1

for i in sOriginalStationsOwn:
    index_row = index_row+1
    ws.cell(row=index_row, column = 1, value = scenario)
    ws.cell(row=index_row, column = 2, value = i)
    if i in sOriginalStationsPotential:
        ws.cell(row=index_row, column = 3, value = ovLocate[i])
        ws.cell(row=index_row, column = 7, value = pLocationCost[i])
    else:
        ws.cell(row=index_row, column = 3, value = 0)
        ws.cell(row=index_row, column = 7, value = 0)
    ws.cell(row=index_row, column = 4, value = ovQuantityUnitsCapacity[i])
    ws.cell(row=index_row, column = 5, value = pStationCapacity[i])
    ws.cell(row=index_row, column = 6, value = pStationUnitCapacity[i])
    ws.cell(row=index_row, column = 8, value = pCostUnitCapacity[i])

oTotalRefuellingCost = sum(pQuantityVehicles[v,p]*(ovRefuelQuantity[i,v,p]*pPrice[i] + (pOpportunityCost[v] + 2*pVariableCost[v]*pDistanceOOP[i,p])*ovRefuel[i,v,p])  
                                   for (i,v,p) in sStationsVehiclesPaths)
oTotalLocationCost = sum(pLocationCost[i]*ovLocate[i] for i in sOriginalStationsPotential) + sum(pCostUnitCapacity[i]*ovQuantityUnitsCapacity[i] for i in sOriginalStationsOwn)
oTotalDiscount = sum(pDiscount[l,g]*ovQuantityPurchasedRange[l,g] for (l,g) in sSuppliersRanges)
oTotalCost = oTotalRefuellingCost + oTotalLocationCost - oTotalDiscount

sheet = 'oTotalCost'
sh_remove = workbook[sheet]
workbook.remove(sh_remove)
ws = workbook.create_sheet(sheet)

oColumnsTotalCost = ["Scenario", "TotalRefuellingCost", "TotalLocationCost", "TotalDiscount", "TotalCost"]
write_column_titles(ws, oColumnsTotalCost)

ws.cell(row=2, column=1, value=scenario)
ws.cell(row=2, column=2, value=oTotalRefuellingCost)
ws.cell(row=2, column=3, value=oTotalLocationCost)
ws.cell(row=2, column=4, value=oTotalDiscount)
ws.cell(row=2, column=5, value=oTotalCost)

workbook.save(output_file)

m.write("RFEP.lp")
m.write("RFEP.sol")

end_time = time.time()

total_time = end_time - start_time
